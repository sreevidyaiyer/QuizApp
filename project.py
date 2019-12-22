# import openpyxl and tkinter modules 
from openpyxl import *
from tkinter import *
# globally declare wb and sheet variable 
# opening the existing excel file 
wb = load_workbook('excel.xlsx')
# create the sheet object 
sheet = wb.active
ans=[0,0,0,0,0,0]
cans=[]
Home1 = Tk()
v = IntVar()

def Leader():
     global Leader
     Score.withdraw()
     Leader = Toplevel()
     Leader.title("Quiz")
     width = 600
     height = 770
     screen_width = Leader.winfo_screenwidth()
     screen_height = Leader.winfo_screenheight()
     x = (screen_width/2) - (width/2)
     y = (screen_height/2) - (height/2)
     Leader.geometry("%dx%d+%d+%d" % (width, height, x, y))
     Leader.configure(background="#a1dbcd")
     Leader.resizable(0, 0)
     photo=PhotoImage(file="g1.gif")
     w=Label(Leader,image=photo)
     w.place(x=0,y=0,width=600,height=100)
     ab=load_workbook('Score.xlsx')
     
     sheet2=ab.active
     current_row2 = sheet2.max_row 
     current_column2 = sheet2.max_column
     sub_li= [[sheet2.cell(row=i,column=1).value, sheet2.cell(row=i,column=3).value,sheet2.cell(row=i,column=2).value] for i in range(2,sheet2.max_row+1)]


     def Sort(sub_li):
         l = len(sub_li) 
         for i in range(0, l): 
             for j in range(0, l-i-1): 
                 if (sub_li[j][1] < sub_li[j + 1][1]): 
                     tempo = sub_li[j] 
                     sub_li[j]= sub_li[j + 1] 
                     sub_li[j + 1]= tempo 
         return sub_li
        
     Sort(sub_li)
     Label(Leader, text="Leader Board", pady=5, font=('Times',30,'bold') , bg="#a1dbcd").place(x=200,y=100,height=40)
     x1=150
     Label(Leader, text="User ID", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=50,y=x1+20,height=20)
     Label(Leader, text="Score", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=350,y=x1+20,height=20)
     Label(Leader, text="User Name", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=180,y=x1+20,height=20)
     x1=x1+40
     for i in sub_li:
         Label(Leader, text=i[0], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=50,y=x1+20)
         Label(Leader, text=i[1], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=350,y=x1+20)
         Label(Leader, text=i[2], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=180,y=x1+20)
         x1=x1+40
     root.mainloop()       
     
def Leader1():
     global Leader
     Leader = Toplevel()
     Leader.title("Quiz")
     width = 600
     height = 770
     screen_width = Leader.winfo_screenwidth()
     screen_height = Leader.winfo_screenheight()
     x = (screen_width/2) - (width/2)
     y = (screen_height/2) - (height/2)
     Leader.geometry("%dx%d+%d+%d" % (width, height, x, y))
     Leader.configure(background="#a1dbcd")
     photo=PhotoImage(file="g1.gif")
     w=Label(Leader,image=photo)
     Leader.resizable(0, 0)
     w.place(x=0,y=0,width=600,height=100)
     ab=load_workbook('Score.xlsx')
   
     sheet2=ab.active
     current_row2 = sheet2.max_row 
     current_column2 = sheet2.max_column
     sub_li= [[sheet2.cell(row=i,column=1).value, sheet2.cell(row=i,column=3).value,sheet2.cell(row=i,column=2).value] for i in range(2,sheet2.max_row+1)]
     
     def Sort(sub_li):
         l = len(sub_li) 
         for i in range(0, l): 
             for j in range(0, l-i-1): 
                 if (sub_li[j][1] > sub_li[j + 1][1]): 
                     tempo = sub_li[j] 
                     sub_li[j]= sub_li[j + 1] 
                     sub_li[j + 1]= tempo 
         return sub_li
    
     Label(Leader, text="Leader Board", pady=5, font=('Times',30,'bold') , bg="#a1dbcd").place(x=200,y=100,height=40)
     x1=150
     Label(Leader, text="User ID", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=50,y=x1+20,height=20)
     Label(Leader, text="Score", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=350,y=x1+20,height=20)
     Label(Leader, text="User Name", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=180,y=x1+20,height=20)
     x1=x1+40
     for i in sub_li:
         Label(Leader, text=i[0], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=50,y=x1+20)
         Label(Leader, text=i[1], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=350,y=x1+20)
         Label(Leader, text=i[2], pady=5, font=('Times',15,'bold') , bg="#a1dbcd").place(x=180,y=x1+20)
         x1=x1+40
     root1.mainloop()                
def QuizWindow():
        global Quiz
        qs=load_workbook('Questions.xlsx')
        sheet1=qs.active
        sheet1.cell(row=1, column=1).value = "QUESTION"
        sheet1.cell(row=1, column=2).value = "OPTION 1"
        sheet1.cell(row=1, column=3).value = "OPTION 2"
        sheet1.cell(row=1, column=4).value = "OPTION 3"
        sheet1.cell(row=1, column=5).value = "OPTION 4"
        sheet1.cell(row=1, column=6).value = "CORRECT OPTION NO"
        # Function to take data from GUI 
        # window and write to an excel file
        current_row1 = sheet1.max_row 
        current_column1 = sheet1.max_column 
        root1.withdraw()
        Quiz = Toplevel()
        Quiz.title("Admin Page")
        width = 600
        height = 560
        screen_width = Quiz.winfo_screenwidth()
        screen_height = Quiz.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        Quiz.geometry("%dx%d+%d+%d" % (width, height, x, y))
        Quiz.configure(background="#a1dbcd")
        Quiz.resizable(0, 0)

        #==============================FRAMES=========================================
        Top = Frame(Quiz, bd=2,  relief=RIDGE)
        Top.grid(row=0, column=1)

        #==============================VARIABLES======================================
        Q1= StringVar()
        O1 = StringVar()
        O2 = StringVar()
        O3 = StringVar()
        O4 = StringVar()
        O5 = StringVar()
        #==============================LABELS=========================================
        q1 = Label(Top, text = "ADMIN Login", font=('arial', 15))
        q1.grid(row=0,column=1)
        q = Label(Quiz, text = "Enter new Question:", font=('arial', 14), bd=15 , bg="#a1dbcd")
        q.grid(row=3, sticky="e")
        op1 = Label(Quiz, text = "Enter Option 1:", font=('arial', 14), bd=15 , bg="#a1dbcd")
        op1.grid(row=4, sticky="e")
        op2 = Label(Quiz, text = "Enter Option 2:", font=('arial', 14), bd=15, bg="#a1dbcd")
        op2.grid(row=5, sticky="e")
        op3 = Label(Quiz, text = "Enter Option 3:", font=('arial', 14), bd=15, bg="#a1dbcd")
        op3.grid(row=6, sticky="e")
        op4 = Label(Quiz, text = "Enter Option 4:", font=('arial', 14), bd=15, bg="#a1dbcd")
        op4.grid(row=7, sticky="e")
        op5 = Label(Quiz, text = "Enter Correct Option:", font=('arial', 14), bd=15, bg="#a1dbcd")
        op5.grid(row=8, sticky="e")
        lbl_text = Label(Quiz, font=('arial', 14), bd=15 ,bg="#a1dbcd")
        lbl_text.grid(row=9, columnspan=3)
        #==============================ENTRY WIDGETS==================================
        q_1 = Entry(Quiz, textvariable=Q1, font=(14))
        q_1.grid(row=3, column=1,columnspan=5)
        o_1 = Entry(Quiz, textvariable=O1 , font=(14))
        o_1.grid(row=4, column=1)
        o_2 = Entry(Quiz, textvariable=O2 , font=(14))
        o_2.grid(row=5, column=1)
        o_3 = Entry(Quiz, textvariable=O3 , font=(14))
        o_3.grid(row=6, column=1)
        o_4 = Entry(Quiz, textvariable=O4 , font=(14))
        o_4.grid(row=7, column=1)
        o_5 = Entry(Quiz, textvariable=O5 , font=(14))
        o_5.grid(row=8, column=1)

        

        def excel1(): 

                # resize the width of columns in 
                # excel spreadsheet 
                sheet1.column_dimensions['A'].width = 100
                sheet1.column_dimensions['B'].width = 40
                sheet1.column_dimensions['C'].width = 40
                sheet1.column_dimensions['D'].width = 40
                sheet1.column_dimensions['E'].width = 40
                sheet1.column_dimensions['F'].width = 40
                # write given data to an excel spreadsheet 
                # at particular location 

        excel1()
        def insert():
                # if user does not fill any entry 
                # then print "empty input"
                if (Q1.get() == ""):
                    print("empty input") 
                else:
                    # assigning the max row and max column 
                    # value upto which data is written 
                    # in an excel sheet to the variable 
                    current_row1 = sheet1.max_row 
                    current_column1 = sheet1.max_column 
                    # get method returns current text 
                    # as string which we write into 
                    # excel spreadsheet at particular location 
                    sheet1.cell(row=current_row1 + 1, column=1).value = Q1.get() 
                    sheet1.cell(row=current_row1+ 1, column=2).value = O1.get() 
                    sheet1.cell(row=current_row1+ 1, column=3).value = O2.get()
                    sheet1.cell(row=current_row1+ 1, column=4).value = O3.get()
                    sheet1.cell(row=current_row1+ 1, column=5).value = O4.get()
                    sheet1.cell(row=current_row1+ 1, column=6).value = O5.get()
                    # save the file
                    
                    qs.save('Questions.xlsx')
        def submit1():
                p=0
                if Q1.get() != "" and O1.get() != "" and O2.get()!="":
                  for i in range(1,sheet.max_row+1):
                      cell_obj1=sheet1.cell(row=i,column=1)
                      cell_obj2=sheet1.cell(row=i,column=2)
                      cell_obj3=sheet1.cell(row=i,column=3)
                      cell_obj4=sheet1.cell(row=i,column=4)
                      cell_obj5=sheet1.cell(row=i,column=5)
                      cell_obj6=sheet1.cell(row=i,column=6)
                      if(Q1.get()==cell_obj1.value and O1.get()==cell_obj2.value and O2.get()==cell_obj3.value and O3.get()==cell_obj4.value and O4.get()==cell_obj5.value and O5.get()==cell_obj6.value):
                          p=p+1					
                                                
                  if p!=0:
                      lbl_text.config(text="")
                      Q1.set("")
                      O1.set("")
                      O2.set("")
                      O3.set("")
                      O4.set("")
                      O5.set("")
                      lbl_text.config(text="You have already registered " , fg="red")
                  else:
                      insert()
                      Q1.set("")
                      O1.set("")
                      O2.set("")
                      O3.set("")
                      O4.set("")
                      O5.set("")
                      print("inserted successfully")
                      lbl_text.config(text="")

                elif Q1.get() == "" or O1.get() == "" or O2.get()=="" :
                  lbl_text.config(text="Please complete the required field!", fg="red")

        #==============================BUTTON WIDGETS=================================

        btn= Button(Quiz, text="Submit", width=20 ,command=submit1)
        btn.grid(pady=25, row=10, columnspan=2)
        btn.bind('<Return>',submit1)
        btn1= Button(Quiz, text="Leader Board", width=20 ,command=Leader1)
        btn1.grid(pady=25, row=11, columnspan=2)
        btn1.bind('<Return>',Leader1 )          
def Score(count,T1,T2):
                  
                  global Score
                  Home2.withdraw()
                  Score = Toplevel()
                  Score.title("Quiz")
                  width = 600
                  height = 770
                  screen_width = Score.winfo_screenwidth()
                  screen_height = Score.winfo_screenheight()
                  x = (screen_width/2) - (width/2)
                  y = (screen_height/2) - (height/2)
                  Score.geometry("%dx%d+%d+%d" % (width, height, x, y))
                  Score.configure(background="#a1dbcd")
                  photo=PhotoImage(file="g1.gif")
                  w=Label(Score,image=photo)
                  x1=100
                  w.place(x=0,y=0,width=600,height=100)
                  ab=load_workbook('Score.xlsx')
                  sheet2=ab.active
                  sheet2.column_dimensions['A'].width = 40
                  sheet2.column_dimensions['B'].width = 40
                  sheet2.column_dimensions['C'].width = 40
                  sheet2.cell(row=1, column=1).value = "ID"
                  sheet2.cell(row=1, column=2).value = "USERNAME"
                  sheet2.cell(row=1, column=3).value = "SCORE"
                  current_row2 = sheet2.max_row 
                  current_column2 = sheet2.max_column 
                  # get method returns current text 
                  # as string which we write into 
                  # excel spreadsheet at particular location
                  p1=0
                  for i in range(1,sheet2.max_row+1):
                      cell_obj1=sheet2.cell(row=i,column=1)
                      cell_obj2=sheet2.cell(row=i,column=2)
                      cell_obj3=sheet2.cell(row=i,column=3)
                      if T2==cell_obj1.value:
                          p1=p1+1
                          sheet2.cell(row=i, column=3).value =count

                  if i==sheet2.max_row and p1==0:
                      sheet2.cell(row=current_row2 + 1, column=1).value =T2
                      sheet2.cell(row=current_row2+ 1, column=2).value = T1
                      sheet2.cell(row=current_row2+ 1, column=3).value = count
                      
                  Score.resizable(0, 0)
                  ab.save('Score.xlsx')
                  Label(Score, text="Score Card", pady=5, font=('Times',30,'bold') , bg="#a1dbcd").place(x=200,y=x1+50,height=40)
                  Label(Score, text="UserName:", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=50,y=x1+125,height=40)
                  Label(Score, text=T1, pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=190,y=x1+125,height=40)
                  Label(Score, text="User Id:", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=50,y=x1+185,height=40)
                  Label(Score, text=T2, pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=180,y=x1+185,height=40)
                  Label(Score, text="Score :", pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=50,y=x1+235,height=40)
                  Label(Score, text=count, pady=5, font=('Times',20,'bold') , bg="#a1dbcd").place(x=180,y=x1+235,height=40)
    

                  b1=Button(Score, text="Leader Board", width=45 ,fg="blue",bg="pink",command=Leader)
                  b1.place(x=180,y=500,width=100,height=50)  
                  root.mainloop()

def HomeWindow2(T1,T2):
        global Home2
        Home1.withdraw()
        Home2 = Toplevel()
        Home2.title("Quiz")
        width = 600
        height = 770
        screen_width = Home2.winfo_screenwidth()
        screen_height = Home2.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        Home2.geometry("%dx%d+%d+%d" % (width, height, x, y))
        Home2.configure(background="#a1dbcd")
        var1 = StringVar()
        var2 = StringVar()
        var3 = StringVar()
        photo=PhotoImage(file="g1.gif")
        w=Label(Home2,image=photo)
        w.place(x=0,y=0,width=600,height=100)
        k=130
        qs=load_workbook('Questions.xlsx')
        sheet1=qs.active
        l=110
        f=-1
        Home2.resizable(0, 0)
        for i in range(5,sheet1.max_row+1):
                f=f+1
                cell_obj1=sheet1.cell(row=i,column=1)
                cell_obj2=sheet1.cell(row=i,column=2)
                cell_obj3=sheet1.cell(row=i,column=3)
                cell_obj4=sheet1.cell(row=i,column=4)
                cell_obj5=sheet1.cell(row=i,column=5)
                cell_obj6=sheet1.cell(row=i,column=6)
                d=cell_obj6.value
                cans.append(d)
                l2=Label(Home2, text="Q)"+cell_obj1.value,font=('Times',13,'bold'),fg='black' , bg="#a1dbcd")
                l2.place(x=1 ,y=l ,height=75)
                Label(Home2, text="a)"+cell_obj2.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+25,height=20)
   
                Label(Home2, text="b)"+cell_obj3.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+50,height=20)
          
                Label(Home2, text="c)"+cell_obj4.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+75,height=20)
            
                Label(Home2, text="d)"+cell_obj5.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+100,height=20)
                Label(Home2, text="Enter Correct Option :", pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=i,y=k+125,height=20)
                l=l+185
                k=k+200
        if i!=7:
             cans.append(5)  
        e1 = Entry(Home2,textvariable=var1)
        e1.place(x=200,y=255,height=20,width=100)
        e2 = Entry(Home2,textvariable=var2)
        e2.place(x=200,y=455,height=20,width=100)
        temp=0
        if i==7:
              e3 = Entry(Home2,textvariable=var3)
              e3.place(x=200,y=655,height=20,width=100)
              temp=1
        def Choice():
            count =0
            ans[3]=e1.get()
            ans[4]=e2.get()
            if temp==1:
                 ans[5]=e3.get()
            else:
                 ans[5]=0                        
            for i in range(0,6):
                
                if ans[i]==cans[i]:
                    
                    count=count+1

            Score(count,T1,T2)
        b1=Button(Home2, text="Submit", width=45 ,fg="blue",bg="pink",command=Choice)
        b1.place(x=400,y=620,width=100,height=50)        
        root.mainloop()
def HomeWindow(T1,T2):
        global Home
        root.withdraw()
        Home = Toplevel()
        Home.title("Quiz")
        width = 600
        height = 770
        screen_width = Home.winfo_screenwidth()
        screen_height = Home.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        Home.geometry("%dx%d+%d+%d" % (width, height, x, y))
        Home.configure(background="#a1dbcd")
        var1 = StringVar()
        var2 = StringVar()
        var3 = StringVar()
        photo=PhotoImage(file="g1.gif")
        w=Label(Home,image=photo)
        w.place(x=0,y=0,width=600,height=100)
        k=130
        qs=load_workbook('Questions.xlsx')
        sheet1=qs.active
        l=110
        f=-1
        Home.resizable(0, 0)
        for i in range(2,5):
                f=f+1
                cell_obj1=sheet1.cell(row=i,column=1)
                cell_obj2=sheet1.cell(row=i,column=2)
                cell_obj3=sheet1.cell(row=i,column=3)
                cell_obj4=sheet1.cell(row=i,column=4)
                cell_obj5=sheet1.cell(row=i,column=5)
                cell_obj6=sheet1.cell(row=i,column=6)
                d=cell_obj6.value
                cans.append(d)
                l2=Label(Home, text="Q)"+cell_obj1.value,font=('Times',13,'bold'),fg='black' , bg="#a1dbcd")
                l2.place(x=1 ,y=l ,height=55)
                Label(Home, text="a)"+cell_obj2.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+25,height=20)
   
                Label(Home, text="b)"+cell_obj3.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+50,height=20)
          
                Label(Home, text="c)"+cell_obj4.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+75,height=20)
            
                Label(Home, text="d)"+cell_obj5.value, pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=1,y=k+100,height=20)
                Label(Home, text="Enter Correct Option :", pady=5, font=('Times',10,'bold') , bg="#a1dbcd").place(x=i,y=k+125,height=20)
                l=l+185
                k=k+200
        e1 = Entry(Home,textvariable=var1)
        e1.place(x=200,y=255,height=20,width=100)
        e2 = Entry(Home,textvariable=var2)
        e2.place(x=200,y=455,height=20,width=100)
        e3 = Entry(Home,textvariable=var3)
        e3.place(x=200,y=655,height=20,width=100)
        def Choice1():
                ans[0]=e1.get()
                ans[1]=e2.get()
                ans[2]=e3.get()
             
                HomeWindow2(T1,T2)
        b1=Button(Home, text="Next", width=45 ,fg="blue",bg="pink",command=Choice1)
        b1.place(x=400,y=620,width=100,height=50)        
        root.mainloop()
def Back():
    Home.destroy()
    root.deiconify()

def excel(): 
	
	# resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 40
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 20
	
        # write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "ID"
	sheet.cell(row=1, column=2).value = "USERNAME"
	sheet.cell(row=1, column=3).value = "PASSWORD"
	sheet.cell(row=1, column=4).value = "EMAIL"
# Function to take data from GUI 
# window and write to an excel file
current_row = sheet.max_row 
current_column = sheet.max_column 
def Admin():
    global root1
    global Form1
    Home1.withdraw()
    root1 = Toplevel()
    root1.title("ADMIN Login")
    width = 600
    height = 480
    screen_width = root1.winfo_screenwidth()
    screen_height = root1.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    root1.geometry("%dx%d+%d+%d" % (width, height, x, y))
    root1.resizable(0, 0)
    root1.configure(background="#a1dbcd")
    #==============================VARIABLES======================================
    USERNAME1 = StringVar()
    PASSWORD1 = StringVar()
    #==============================FRAMES=========================================
    Top = Frame(root1, bd=2,  relief=RIDGE)
    Top.pack(side=TOP, fill=X)
    Form1 = Frame(root1, height=400, bg="#a1dbcd")
    Form1.pack(side=TOP, pady=1)

    #==============================LABELS=========================================
    lbl_title = Label(Top, text = "ADMIN Login", font=('arial', 15), bg="#a1dbcd")
    lbl_title.pack(fill=X)
    lbl_username = Label(Form1, text = "Username:", font=('arial', 14), bd=15, bg="#a1dbcd")
    lbl_username.grid(row=0, sticky="e")
    lbl_password = Label(Form1, text = "Password:", font=('arial', 14), bd=15, bg="#a1dbcd")
    lbl_password.grid(row=1, sticky="e")
    #==============================ENTRY WIDGETS==================================
    user_name = Entry(Form1, textvariable=USERNAME1, font=(14))
    user_name.grid(row=0, column=1)
    password = Entry(Form1, textvariable=PASSWORD1, show="*", font=(14))
    password.grid(row=1, column=1)

    def AD_Login():
      lbl_text = Label(Form1)
      lbl_text.grid(row=5, columnspan=4)
      if USERNAME1.get() == "" or PASSWORD1.get() == "" :
          lbl_text.config(text="Please complete the required field!", fg="red", bg="#a1dbcd")
      
      if USERNAME1.get() != "" or PASSWORD1.get() != "" :
           
              lbl_text.config(text=" ", bg="#a1dbcd")
              if(USERNAME1.get()=="Admin" and PASSWORD1.get()=="1234ABCD"):
                  lbl_text.config(text="Successfully Login")
                  QuizWindow()
                  return
              else:
                  lbl_text.config(text="   Invalid username or password   ", fg="red", bg="#a1dbcd")
                  USERNAME1.set("")
                  PASSWORD1.set("")
      


    #==============================BUTTON WIDGETS=================================

    btn = Button(Form1, text="Login", width=45 ,command=AD_Login)
    btn.grid(pady=25, row=6, columnspan=3)
    btn.bind('<Return>', AD_Login)
    
def func():
        global root
        global Form
        Home1.withdraw()
        root = Toplevel()
        root.title("Python: Simple Login Application")
        width = 600
        height = 480
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width/2) - (width/2)
        y = (screen_height/2) - (height/2)
        root.geometry("%dx%d+%d+%d" % (width, height, x, y))
        root.resizable(0, 0)
        root.configure(background="#a1dbcd")
        #==============================VARIABLES======================================
        USERNAME = StringVar()
        PASSWORD = StringVar()

        #==============================FRAMES=========================================
        Top = Frame(root, bd=2,  relief=RIDGE)
        Top.pack(side=TOP, fill=X)
        Form = Frame(root, height=400, bg="#a1dbcd")
        Form.pack(side=TOP, pady=1)

        #==============================LABELS=========================================
        lbl_title = Label(Top, text = "User Login", font=('arial', 15), bg="#a1dbcd")
        lbl_title.pack(fill=X)
        lbl_username = Label(Form, text = "Username:", font=('arial', 14), bd=15, bg="#a1dbcd")
        lbl_username.grid(row=0, sticky="e")
        lbl_password = Label(Form, text = "Password:", font=('arial', 14), bd=15, bg="#a1dbcd")
        lbl_password.grid(row=1, sticky="e")
        lbl_text = Label(Form)
        lbl_text.grid(row=5, columnspan=4)

        #==============================ENTRY WIDGETS==================================
        user_name = Entry(Form, textvariable=USERNAME, font=(14))
        user_name.grid(row=0, column=1)
        password = Entry(Form, textvariable=PASSWORD, show="*", font=(14))
        password.grid(row=1, column=1)

        excel()

        

        def Login():
            lbl_text = Label(Form)
            lbl_text.grid(row=5, columnspan=4)
            if USERNAME.get() == "" or PASSWORD.get() == "" :
                lbl_text.config(text="Please complete the required field!", fg="red", bg="#a1dbcd")
            if USERNAME.get() != "" or PASSWORD.get() != "" :
                # opening the existing excel file 
                wb = load_workbook('excel.xlsx') 

                # create the sheet object 
                sheet = wb.active

                for i in range(1,sheet.max_row+1):
                    cell_obj1=sheet.cell(row=i,column=2)
                    cell_obj2=sheet.cell(row=i,column=3)
                    cell_obj4=sheet.cell(row=i,column=1)
                    if(USERNAME.get()==cell_obj1.value and PASSWORD.get()==cell_obj2.value):
                        T1=cell_obj1.value
                        T2=cell_obj4.value
                        
                        HomeWindow(T1,T2)
                        USERNAME.set("")
                        PASSWORD.set("")
                        lbl_text.config(text="", bg="#a1dbcd")
                        return
                lbl_text.config(text="Invalid username or password", fg="red", bg="#a1dbcd")
                USERNAME.set("")
                PASSWORD.set("")
        def Register():

          def insert():
            # if user does not fill any entry 
            # then print "empty input"
            if (USERNAME.get() == "" and EMAIL.get() == "" and PASSWORD.get() == ""):
                print("empty input") 
            else:
                # assigning the max row and max column 
                # value upto which data is written 
                # in an excel sheet to the variable 
                current_row = sheet.max_row 
                current_column = sheet.max_column 
                # get method returns current text 
                # as string which we write into 
                # excel spreadsheet at particular location 
                T2=sheet.cell(row=current_row + 1, column=1).value = ID.get() 
                T1=sheet.cell(row=current_row + 1, column=2).value = USERNAME.get() 
                sheet.cell(row=current_row + 1, column=3).value = PASSWORD.get()
                sheet.cell(row=current_row + 1, column=4).value = EMAIL.get()
                # save the file 
                wb.save('excel.xlsx')
                        
          def submit():
              p=0
              s1=EMAIL.get()
              if USERNAME.get() != "" and PASSWORD.get() != "" and EMAIL.get()!="":
                  for i in range(1,sheet.max_row+1):
                      cell_obj1=sheet.cell(row=i,column=2)
                      cell_obj2=sheet.cell(row=i,column=4)
                      cell_obj3=sheet.cell(row=i,column=3)
                      cell_obj4=sheet.cell(row=i,column=1)
                      if(USERNAME.get()==cell_obj1.value and PASSWORD.get()==cell_obj3.value and EMAIL.get()==cell_obj2.value and ID.get()==cell_obj4.value):
                          p=p+1					
                                                
                  if p!=0:
                      lbl_text.config(text="", bg="#a1dbcd")
                      USERNAME.set("")
                      PASSWORD.set("")
                      EMAIL.set("")
                      lbl_text.config(text="You have already registered " , fg="red", bg="#a1dbcd")
                  else:
                      if re.findall('\S+@\S+',s1):
                          T1=USERNAME.get()
                          T2=ID.get()
                          insert()
                          HomeWindow(T1,T2)
                          lbl_text.config(text="", bg="#a1dbcd")
                      else:
                          lbl_text.config(text="Please enter correct email address!", fg="red", bg="#a1dbcd")

              elif USERNAME.get() == "" or PASSWORD.get() == "" or EMAIL.get()=="" :
                  lbl_text.config(text="Please complete the required field!", fg="red", bg="#a1dbcd")
          EMAIL=StringVar()
          ID= StringVar()
          lbl_password = Label(Form, text = "Email:", font=('arial', 14), bd=15, bg="#a1dbcd")
          lbl_password.grid(row=2, sticky="e")
          password = Entry(Form, textvariable=EMAIL , font=(14))
          password.grid(row=2, column=1)
          lbl_password = Label(Form, text = "ID:", font=('arial', 14), bd=15, bg="#a1dbcd")
          lbl_password.grid(row=3, sticky="e")
          password = Entry(Form, textvariable=ID , font=(14))
          password.grid(row=3, column=1)
          lbl_text = Label(Form)
          lbl_text.grid(row=5, columnspan=4)
          btn_register = Button(Form, text="Submit", width=45 ,command=submit)
          btn_register.grid(pady=25, row=8, columnspan=5)
          btn_register.bind('<Return>', submit)

        #==============================BUTTON WIDGETS=================================

        btn_register = Button(Form, text="Login", width=45 ,command=Login)
        btn_register.grid(pady=25, row=6, columnspan=3)
        btn_register.bind('<Return>', Login)
        btn_register = Button(Form, text="Register", width=45 ,command=Register)
        btn_register.grid(pady=25, row=7, columnspan=4)
        btn_register.bind('<Return>', Register)



   
# Driver code
if __name__ == "__main__":
    Home1.title("Quiz")
    width = 600
    height = 560
    screen_width = Home1.winfo_screenwidth()
    screen_height = Home1.winfo_screenheight()
    x = (screen_width/2) - (width/2)
    y = (screen_height/2) - (height/2)
    Home1.geometry("%dx%d+%d+%d" % (width, height, x, y))
    Home1.configure(background="#a1dbcd")

    #==============================BUTTON WIDGETS=================================

    btn1 = Button(Home1, text="ADMIN", width=45 ,command=Admin)
    btn1.grid(pady=25, row=6, columnspan=3)
    btn1.bind('<Return>',Admin)
    btn2 = Button(Home1, text="STUDENT", width=45 ,command=func)
    btn2.grid(pady=25, row=7, columnspan=4)
    btn2.bind('<Return>',func)
    
    
    # start the GUI
    Home1.mainloop()	
